import os
import pandas as pd
import lkml
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def main():
    root_folder = 'ONELooker'
    directory = 'C:/Users/NC/Documents/ONELooker/LOOKML_one_bkg_doc_spoke/'
    results = []

    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.endswith('.view.lkml'):
                full_file_path = os.path.join(root, file)
                relative_folder = root.split(root_folder)[-1]
                if relative_folder == '':
                    folder = ''
                else:
                    folder = "/" + relative_folder.lstrip('/').replace('\\', '/') + "/"
                with open(full_file_path, 'r') as f:
                    parsed_content = lkml.load(f)

                for view in parsed_content.get('views', []):
                    view_name = view['name']
                    for dim_group in view.get('dimension_groups', []):
                        
                        dim_group_name = dim_group['name']
                        dim_group_body = dim_group  # Modify this if you want specific parts of the body.
                        convert_tz_param = dim_group.get('convert_tz')
                        print(convert_tz_param)

                        if convert_tz_param != 'no':
                            results.append((folder, file, view_name, dim_group_name))

    df = pd.DataFrame(results, columns=["Folder", "Filename", "View Name", "Dimension Group"])
    export_to_excel(df, "Test20.xlsx")

def export_to_excel(df, filename):
    workbook = Workbook()
    sheet = workbook.active

    for col_num, header in enumerate(df.columns, 1):
        column_letter = get_column_letter(col_num)
        sheet[f"{column_letter}1"] = header

    for row_num, row in enumerate(df.values, 2):
        for col_num, value in enumerate(row, 1):
            column_letter = get_column_letter(col_num)
            sheet[f"{column_letter}{row_num}"] = value

    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    workbook.save(filename)

if __name__ == "__main__":
    main()
